from flask import Flask, request, jsonify, send_file, render_template
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__, static_folder="static", template_folder="templates")

EXCEL_FILE = os.path.join(os.getcwd(), "planning_arcabo.xlsx")

def save_to_excel(data):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "planning_arcabo"
        ws.append([
            "Startdatum", "Einddatum", "Code", "Locatie",
            "Model", "Lengte", "Breedte", "Zagerij",
            "Dynamisch", "Statisch"
        ])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        data.get("startdatum", ""),
        data.get("einddatum", ""),
        data.get("code", ""),
        data.get("locatie", ""),
        data.get("model", ""),
        data.get("lengte", ""),
        data.get("breedte", ""),
        data.get("zagerij", ""),
        data.get("dynamisch", ""),
        data.get("statisch", "")
    ])
    wb.save(EXCEL_FILE)

@app.route("/save", methods=["POST"])
def save():
    data = request.get_json()
    if not data:
        return jsonify({"error": "❌ Geen geldige JSON ontvangen."}), 400
    save_to_excel(data)
    return jsonify({"message": "✅ Opdracht succesvol opgeslagen in Excel!"})

@app.route("/")
def home():
    return render_template("plannen.html")

@app.route("/download")
def download():
    if os.path.exists(EXCEL_FILE):
        return send_file(EXCEL_FILE, as_attachment=True)
    return jsonify({"error": "❌ Geen Excel-bestand gevonden."}), 404

if __name__ == "__main__":
    app.run(debug=True)
